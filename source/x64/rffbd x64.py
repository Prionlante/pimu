import sys
import os
import math
from PyQt6 import QtWidgets, QtCore, QtGui
import interfaceX64 as ww
import sqlite3
from openpyxl import load_workbook, Workbook
import numpy as np
from TableModelX64 import TableModel
from statisticsM import correlation, mean
from statisticsM import quantiles, minStat
from iufp import iufp

class MainWindow(QtWidgets.QMainWindow, ww.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.addFileButton.clicked.connect(self.fileDialog)
        self.data = []
        self.filename = ""
        self.corr1Item = ""
        self.corr2Item = ""
        self.minAgeData = -1
        self.maxAgeData = -1
        self.sexData = -1
        self.calculateButton.clicked.connect(self.calculate)
        self.applyAgesButton.clicked.connect(self.yearsFilter)
        self.saveFileButton.clicked.connect(self.save)
        self.numberBox.currentIndexChanged[int].connect(self.onChoose)
        self.corr1Box.currentIndexChanged[int].connect(self.onChooseCorr)
        self.corr2Box.currentIndexChanged[int].connect(self.onChooseCorr)
        self.pushButton.clicked.connect(self.correlation)
        self.openBd.clicked.connect(self.getDB)
        self.statisticButton.clicked.connect(self.getStatistic)


    def getQuery(self, arg1="*", arg2=""):
        minAge = self.minAge.text()
        maxAge = self.maxAge.text()
        query = ""
        if minAge and maxAge:
            if int(minAge) > int(maxAge):
                self.label.setText("Минимальный возраст должен быть\nменьше либо равен максимальному")
            else:
                if self.girlsRadio.isChecked():
                    query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 2 AND lmo <= {maxAge} AND lmo >= {minAge}'
                if self.boysRadio.isChecked():
                    query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 1 AND lmo <= {maxAge} AND lmo >= {minAge}'
                if self.allRadio.isChecked():
                    query = f'SELECT {arg1}{arg2} FROM kids WHERE lmo <= {maxAge} AND lmo >= {minAge}'
                if self.cancelRadio.isChecked():
                    query = f'SELECT {arg1}{arg2} FROM kids WHERE lmo <= {maxAge} AND lmo >= {minAge}'
        elif maxAge:
            if self.girlsRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 2 AND lmo <= {maxAge}'
            if self.boysRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 1 AND lmo <= {maxAge}'
            if self.allRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE lmo <= {maxAge}'
            if self.cancelRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE lmo <= {maxAge}'
        elif minAge:
            if self.girlsRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 2 AND lmo >= {minAge}'
            if self.boysRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 1 AND lmo >= {minAge}'
            if self.allRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE lmo >= {minAge}'
            if self.cancelRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE lmo >= {minAge}'
        else:
            if self.girlsRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 2'
            if self.boysRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids WHERE sex = 1'
            if self.allRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids'
            if self.cancelRadio.isChecked():
                query = f'SELECT {arg1}{arg2} FROM kids'
        return query


    def getStatistic(self):
        db = sqlite3.connect(self.filename)
        cur = db.cursor()
        cur.execute(self.getQuery())
        fetchedData = list(cur.fetchall())
        data = []
        for item in fetchedData:
            data.append(item)

        stats = minStat(data)
        n = stats['dta']['count']
        data = [
            [f"all = {n}", "m", "kv", "mean", "std", "min", "percentille 25", "percentille 50", "percentille 75", "max"]
        ]

        for key in list(stats.keys()):
            data.append([key, f"{round(stats[key]['std']/math.sqrt(n), 5)}", f"{round(stats[key]['std']/stats[key]['mean']*100, 5)}%"]+[f"{round(stats[key][k], 5)}" for k in list(stats[key].keys()) if k!= "count"])
        
        self.setData(data)
        


    def setData(self, data):
        model = TableModel(data)
        self.tableOutputExcel.setModel(model)
        self.tableOutputExcel.show()

    def yearsFilter(self):
        self.numberBox.clear()
        minAge = self.minAge.text()
        maxAge = self.maxAge.text()
        self.minAgeData = minAge
        self.maxAgeData = maxAge
        try:
            db = sqlite3.connect(self.filename)
            cur = db.cursor()
            cur.execute(self.getQuery())
            names = list(map(lambda x: x[0], cur.description))
            for i in range(len(names)):
                names[i] = names[i].upper()
            kids = list(cur.fetchall())
            data = []
            data.append(names)
            for kid in kids:
                data.append(kid)
                self.numberBox.addItem(f'{kid[0]}.{kid[1]}')
            self.data = data
            self.setData(data)
        except:
            data = [[]]
            self.data=data
            self.setData(data)

    def fileDialog(self):
        path = QtWidgets.QFileDialog.getOpenFileName(self, 'Открыть', 'C://Users/{}'.format(os.environ.get('USERNAME')), 'Excel Files (*.xlsx)')
        if path != ('', ''):
            self.fileNameEdit.setText(path[0])
    
    def getDB(self):
        path = QtWidgets.QFileDialog.getOpenFileName(self, 'Открыть', 'C://Users/{}'.format(os.environ.get('USERNAME')), 'Database Files (*.db)')
        if path != ('', ''):
            self.filename = path[0]
            self.calculate()

    def radioSelected(self):
        rb = self.sender()
        self.numberBox.clear()
        if rb.isChecked():
            try:
                data = [[]]
                db = sqlite3.connect(self.filename)
                cur = db.cursor()
                if rb.text() == 'Мальчики':
                    self.sexData = 1
                    cur.execute('SELECT * FROM kids WHERE sex = 1')
                    names = list(map(lambda x: x[0], cur.description))
                    for i in range(len(names)):
                        names[i] = names[i].upper()
                    boys = list(cur.fetchall())
                    data = []
                    data.append(names)
                    for boy in boys:
                        data.append(boy)
                        self.numberBox.addItem(f'{boy[0]}.{boy[1]}')
                elif rb.text() == 'Девочки':
                    self.sexData = 2
                    cur.execute('SELECT * FROM kids WHERE sex = 2')
                    names = list(map(lambda x: x[0], cur.description))
                    for i in range(len(names)):
                        names[i] = names[i].upper()
                    girls = list(cur.fetchall())
                    data = []
                    data.append(names)
                    for girl in girls:
                        data.append(girl)
                        self.numberBox.addItem(f'{girl[0]}.{girl[1]}')
                elif rb.text() == 'Все':
                    cur.execute('SELECT * FROM kids')
                    names = list(map(lambda x: x[0], cur.description))
                    for i in range(len(names)):
                        names[i] = names[i].upper()
                    kids = list(cur.fetchall())
                    data = []
                    data.append(names)
                    for kid in kids:
                        data.append(kid)
                        self.numberBox.addItem(f'{kid[0]}.{kid[1]}')
                else:
                    cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 1')
                    boysCount = cur.fetchall()[0][0]
                    cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 2')
                    girlsCount = cur.fetchall()[0][0]
                    data = [
                        ['База данных', 'успешно создана!'],
                        ['',''],
                        ['Мальчиков:', boysCount],
                        ['Девочек:', girlsCount],
                        ['',''],
                    ]
                self.data = data
                self.setData(data)
                cur.close()
                db.close()
            except:
                self.setData(data)    
    
    def calculate(self):
        self.data = []
        self.label.setText("")
        fs = self.fileNameEdit.text().split('/')
        if self.filename == "":
            filename = './databases/' + fs[fs.__len__()-1] + '.db'
        else:
            filename = self.filename
        try:
            db = sqlite3.connect(filename)
            cur = db.cursor()
            if self.filename != filename:
                wb = load_workbook(self.fileNameEdit.text())
                cur.execute('''CREATE TABLE IF NOT EXISTS kids(
                    id INTEGER PRIMARY KEY,
                    fio TEXT,
                    sex INTEGER,
                    vzrg INTEGER,
                    lmom REAL,
                    lmo REAL,
                    dta REAL,
                    mta REAL,
                    ik2 REAL,
                    ogka REAL,
                    st REAL,
                    gela REAL,
                    gi REAL,
                    sada REAL,
                    dada REAL,
                    pd INTEGER,
                    hss INTEGER,
                    dp REAL,
                    udobse REAL,
                    mok REAL,
                    ifi REAL,
                    vik REAL,
                    via REAL,
                    ir REAL,
                    hdd INTEGER,
                    do REAL,
                    mdo REAL,
                    soma INTEGER,
                    kdp INTEGER,
                    kdl INTEGER, 
                    si REAL,
                    iufp REAL
                )''')
                sheet = wb.active
                data = []
                kids = []
                kfs = []
                for i in range(sheet.max_row):
                    row = []
                    for j in range(sheet.max_column):
                        cell = sheet.cell(row = i+1, column = j+1)
                        row.append(cell.value)
                    data.append(row)
                for i in range(data.__len__()):
                    if i == 0:
                        pass
                    else:
                        fio = data[i][0]
                        sex = data[i][1]
                        years = data[i][2]
                        lmom = data[i][3]
                        lmo = data[i][4]
                        dta = data[i][5]
                        mta = data[i][6]
                        ogka = data[i][7]
                        gela = data[i][8]
                        sada = data[i][9]
                        dada = data[i][10]
                        hss = data[i][11]
                        hdd = data[i][12]
                        soma = data[i][13]
                        kdp = data[i][14]
                        kdl = data[i][15]
                        ik2 = mta/(dta/100*dta/100)
                        st = (100+dta+mta-160)/100
                        gi = gela*1000/mta
                        pd = sada-dada
                        dp = hss/hdd
                        udobse = 90.97+0.54*dp-0.57*dada-0.61*lmo
                        mok = udobse*hss
                        ifi = 0.011*hss+0.014*sada+0.008*dada+0.014*lmo-0.009*(dta+mta)-0.27
                        vik = 1-(dada/hss)
                        via = sada * hss
                        ir = (sada*hss)/100
                        do = gela*1000*17/100
                        mdo = hdd * do / 1000
                        si = kdp/mta
                        kids.append([fio, sex, years, lmom, lmo, dta, mta, ik2, ogka, st, gela, gi, sada, dada, pd, hss, dp, udobse, mok, ifi, vik, via, ir, hdd, do, mdo, soma, kdp, kdl, si])
                        kfs.append([dta, mta, ik2, ogka, st, gela, gi, sada, dada, pd, hss, dp, udobse, mok, ifi, vik, via, ir, hdd, do, mdo, soma, kdp, kdl, si])
                        iufp_res = iufp(
                            [dta, mta, ik2, ogka, st, gela, gi, sada, dada, pd, hss, dp, udobse, mok, ifi, vik, via, ir, hdd, do, mdo, soma, kdp, kdl, si], 
                            quantiles(kfs)
                            )
                        query = f'''
                        INSERT INTO kids (id, fio, sex, vzrg, lmom, lmo, dta, mta, ik2, ogka, st, gela, gi, sada, dada, 
                        pd, hss, dp, udobse, mok, ifi, vik, via, ir, hdd, do, mdo, soma, kdp, kdl, si, iufp)
                        VALUES ({i}, \'{fio}\', {sex}, {years}, {lmom}, {lmo}, {dta}, {mta}, {ik2}, {ogka}, {st}, {gela}, {gi},
                        {sada}, {dada}, {pd}, {hss}, {dp}, {udobse}, {mok}, {ifi}, {vik}, {via}, {ir}, {hdd}, {do},
                        {mdo}, {soma}, {kdp}, {kdl},  {si}, {iufp_res})'''
                        self.numberBox.addItem(f'{i}.{fio}')
                        try:
                            cur.execute(query)
                        except:
                            pass
            cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 1')
            boysCount = cur.fetchall()[0][0]
            cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 2')
            girlsCount = cur.fetchall()[0][0]
            cur.execute('SELECT * FROM kids')
            names = list(map(lambda x: x[0], cur.description))
            for name in names:
                self.corr1Box.addItem(name)
                self.corr2Box.addItem(name)
            data = [
                ['База данных', 'успешно создана!'],
                ['',''],
                ['Мальчиков:', boysCount],
                ['Девочек:', girlsCount],
                ['',''],
            ]
            self.setData(data)
            cur.close()
            db.commit()
            db.close()
        except Exception as e:
            print(e)
            self.label.setText("Неправильное название файла")

    def onChooseCorr(self):
        self.corr1Item = self.corr1Box.currentText()
        self.corr2Item = self.corr2Box.currentText()

    def correlation(self):
        db = sqlite3.connect(self.filename)
        cur = db.cursor()
        cur.execute(self.getQuery(self.corr1Item+", ", self.corr2Item))
        res = cur.fetchall()
        meanData = mean(res)
        print(meanData)
        sp = correlation(res, 'spearman')
        kd = correlation(res, 'kendall')
        ps = correlation(res, 'pearson')
        data = [
            ["Выборка: ", f"{len(res)}", ""],
            ["", "Spearman", ""],
            ["", self.corr1Item, self.corr2Item],
            [self.corr1Item, sp[0][0], sp[0][1]],
            [self.corr2Item, sp[1][0], sp[1][1]],
            ["Rx/y", f"{sp[0][1]*meanData[0]/meanData[1]}", ""],
            ["","",""],
            ["", "Kendall", ""],
            ["", self.corr1Item, self.corr2Item],
            [self.corr1Item, kd[0][0], kd[0][1]],
            [self.corr2Item, kd[1][0], kd[1][1]],
            ["Rx/y", f"{kd[0][1]*meanData[0]/meanData[1]}", ""],
            ["","",""],
            ["", "Pearson", ""],
            ["", self.corr1Item, self.corr2Item],
            [self.corr1Item, ps[0][0], ps[0][1]],
            [self.corr2Item, ps[1][0], ps[1][1]],
            ["Rx/y", f"{ps[0][1]*meanData[0]/meanData[1]}", ""]
        ]
        self.setData(data)

    def onChoose(self, index):
        ind = self.numberBox.currentText().split('.')[0]
        try:
            db = sqlite3.connect(self.filename)
            cur = db.cursor()
            cur.execute(f'SELECT * FROM kids WHERE id = {ind}')
            names = list(map(lambda x: x[0], cur.description))
            for i in range(len(names)):
                names[i] = names[i].upper()
            data = list(cur.fetchall()[0])
            info = dict(zip(names, data))
            table = [
                ['ДАННЫЕ О ПАЦИЕНТЕ', ''],
                ['ФИО', info['FIO']],
                ['Возраст', info['VZRG']],
                ['Пол', 'М' if info['SEX'] == 1 else 'Ж'],
                ['', ''],
                ['АНТРОПОМЕТРИЧЕСКИЕ ПОКАЗАТЕЛИ',''],
                ['Рост', info['DTA']],
                ['Вес', info['MTA']],
                ['LMOM', info['LMOM']],
                ['LMO', info['LMO']],
                ['OGKA', info['OGKA']],
                ['Площадь тела', info['ST']],
                ['GELA', info['GELA']],
                ['SADA', info['SADA']],
                ['DADA', info['DADA']],
                ['Частота сердечных сокращений', info['HSS']],
                ['Частота дыхательных движений', info['HDD']],
                ['SOMA', info['SOMA']],
                ['KDP', info['KDP']],
                ['KDL', info['KDL']],
                ['',''],
                ['ФУНКЦИОНАЛЬНЫЕ ПОКАЗАТЕЛИ', ''],
                ['Индекс Кетла 2', info['IK2']],
                ['Жизненный индекс', info['GI']],
                ['Пульсовое давление', info['PD']],
                ['DP', info['DP']],
                ['UdObSe', info['UDOBSE']],
                ['Минутный объем кровообращения', info['MOK']],
                ['IFI', info['IFI']],
                ['Индекс Кердо', info['VIK']],
                ['Индекс Аболенской', info['VIA']],
                ['Индекс Робинсона', info['IR']],
                ['Дыхательный объем', info['DO']],
                ['Минутный дыхательный объем', info['MDO']],
                ['Силовой индекс', info['SI']],
                ["ИУФП", info['IUFP']]
            ]
            self.data = table
            self.setData(table)
            cur.close()
            db.close()
        except Exception as e:
            print(e)
            data = [[]]
            self.setData(data)

    def save(self):
        path = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить', '', 'Excel Files (*.xlsx)')
        wb = Workbook()
        sheet = wb.active
        i = 1
        for row in self.data:
            j = 1
            for item in row:
                cell = sheet.cell(row = i, column = j)
                cell.value = item
                j = j + 1
            j = 1
            i = i + 1
        wb.save(path[0])
        self.label.setText("Успешно сохранено!")
            
        


def main():
    app = QtWidgets.QApplication(sys.argv)  
    app.setWindowIcon(QtGui.QIcon('../../icon.ico'))
    window = MainWindow()
    window.setWindowIcon(QtGui.QIcon('../../icon.ico'))
    window.show() 
    app.exec()  


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()

