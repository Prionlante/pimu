import sys
import os
from PyQt5 import QtWidgets, QtCore, QtGui
import interfaceX32
import openpyxl # найти библиотеку которая работает с xls файлами
import sqlite3
from openpyxl import load_workbook, Workbook
from TableModel import TableModel

class MainWindow(QtWidgets.QMainWindow, interfaceX32.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.addFileButton.clicked.connect(self.fileDialog)
        self.data = []
        self.filename = ""
        self.calculateButton.clicked.connect(self.calculate)
        self.applyAgesButton.clicked.connect(self.yearsFilter)
        self.saveFileButton.clicked.connect(self.save)
        self.numberBox.currentIndexChanged[int].connect(self.onChoose)
        # self.boysRadio.toggled.connect(self.radioSelected)
        # self.girlsRadio.toggled.connect(self.radioSelected)
        # self.allRadio.toggled.connect(self.radioSelected)
        # self.cancelRadio.toggled.connect(self.radioSelected)

    def setData(self, data):
        model = TableModel(data)
        self.tableOutputExcel.setModel(model)
        self.tableOutputExcel.show()

    def yearsFilter(self):
        self.numberBox.clear()
        minAge = self.minAge.text()
        maxAge = self.maxAge.text()
        try:
            db = sqlite3.connect(self.filename)
            cur = db.cursor()
            if minAge and maxAge:
                if int(minAge) > int(maxAge):
                    self.label.setText("Минимальный возраст должен быть\nменьше либо равен максимальному")
                else:
                    if self.girlsRadio.isChecked():
                        query = f'SELECT * FROM kids WHERE sex = 2 AND vzrg <= {maxAge} AND vzrg >= {minAge}'
                    if self.boysRadio.isChecked():
                        query = f'SELECT * FROM kids WHERE sex = 1 AND vzrg <= {maxAge} AND vzrg >= {minAge}'
                    if self.allRadio.isChecked():
                        query = f'SELECT * FROM kids WHERE vzrg <= {maxAge} AND vzrg >= {minAge}'
                    if self.cancelRadio.isChecked():
                        query = f'SELECT * FROM kids WHERE vzrg <= {maxAge} AND vzrg >= {minAge}'
            elif maxAge:
                if self.girlsRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE sex = 2 AND vzrg <= {maxAge}'
                if self.boysRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE sex = 1 AND vzrg <= {maxAge}'
                if self.allRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE vzrg <= {maxAge}'
                if self.cancelRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE vzrg <= {maxAge}'
            elif minAge:
                if self.girlsRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE sex = 2 AND vzrg >= {minAge}'
                if self.boysRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE sex = 1 AND vzrg >= {minAge}'
                if self.allRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE vzrg >= {minAge}'
                if self.cancelRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE vzrg >= {minAge}'
            else:
                if self.girlsRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE sex = 2'
                if self.boysRadio.isChecked():
                    query = f'SELECT * FROM kids WHERE sex = 1'
                if self.allRadio.isChecked():
                    query = f'SELECT * FROM kids'
                if self.cancelRadio.isChecked():
                    query = f'SELECT * FROM kids'
            cur.execute(query)
            names = list(map(lambda x: x[0], cur.description))
            for i in range(len(names)):
                names[i] = names[i].upper()
            kids = list(cur.fetchall())
            data = []
            data.append(names)
            for kid in kids:
                data.append(kid)
                self.numberBox.addItem(f'{kid[0]}.{kid[1]}')
            self.data = data
            self.setData(data)
        except:
            data = [[]]
            self.data=data
            self.setData(data)

    def fileDialog(self):
        path = QtWidgets.QFileDialog.getOpenFileName(self, 'Открыть', 'C://Users/{}'.format(os.environ.get('USERNAME')), 'Excel Files (*.xlsx)')
        if path != ('', ''):
            self.fileNameEdit.setText(path[0])

    def radioSelected(self):
        rb = self.sender()
        minAge = self.minAge.text()
        maxAge = self.maxAge.text()
        self.numberBox.clear()
        if rb.isChecked():
            try:
                data = [[]]
                db = sqlite3.connect(self.filename)
                cur = db.cursor()
                if rb.text() == 'Мальчики':
                    cur.execute('SELECT * FROM kids WHERE sex = 1')
                    names = list(map(lambda x: x[0], cur.description))
                    for i in range(len(names)):
                        names[i] = names[i].upper()
                    boys = list(cur.fetchall())
                    data = []
                    data.append(names)
                    for boy in boys:
                        data.append(boy)
                        self.numberBox.addItem(f'{boy[0]}.{boy[1]}')
                elif rb.text() == 'Девочки':
                    cur.execute('SELECT * FROM kids WHERE sex = 2')
                    names = list(map(lambda x: x[0], cur.description))
                    for i in range(len(names)):
                        names[i] = names[i].upper()
                    girls = list(cur.fetchall())
                    data = []
                    data.append(names)
                    for girl in girls:
                        data.append(girl)
                        self.numberBox.addItem(f'{girl[0]}.{girl[1]}')
                elif rb.text() == 'Все':
                    cur.execute('SELECT * FROM kids')
                    names = list(map(lambda x: x[0], cur.description))
                    for i in range(len(names)):
                        names[i] = names[i].upper()
                    kids = list(cur.fetchall())
                    data = []
                    data.append(names)
                    for kid in kids:
                        data.append(kid)
                        self.numberBox.addItem(f'{kid[0]}.{kid[1]}')
                else:
                    cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 1')
                    boysCount = cur.fetchall()[0][0]
                    cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 2')
                    girlsCount = cur.fetchall()[0][0]
                    data = [
                        ['База данных', 'успешно создана!'],
                        ['',''],
                        ['Мальчиков:', boysCount],
                        ['Девочек:', girlsCount],
                        ['',''],
                    ]
                self.data = data
                self.setData(data)
                cur.close()
                db.close()
            except:
                self.setData(data)    
    
    def calculate(self):
        self.data = []
        self.label.setText("")
        fs = self.fileNameEdit.text().split('/')
        self.filename = './databases/' + fs[fs.__len__()-1] + '.db'
        try:
            wb = load_workbook(self.fileNameEdit.text())
            db = sqlite3.connect(self.filename)
            cur = db.cursor()
            cur.execute('''CREATE TABLE IF NOT EXISTS kids(
                id INTEGER PRIMARY KEY,
                fio TEXT,
                sex INTEGER,
                vzrg INTEGER,
                lmom REAL,
                lmo REAL,
                dta REAL,
                mta REAL,
                ik2 REAL,
                ogka REAL,
                st REAL,
                gela REAL,
                gi REAL,
                sada REAL,
                dada REAL,
                pd INTEGER,
                hss INTEGER,
                dp REAL,
                udobse REAL,
                mok REAL,
                ifi REAL,
                vik REAL,
                via REAL,
                ir REAL,
                hdd INTEGER,
                do REAL,
                mdo REAL,
                soma INTEGER,
                kdp INTEGER,
                kdl INTEGER, 
                si REAL
            )''')
            sheet = wb.active
            data = []
            for i in range(sheet.max_row):
                row = []
                for j in range(sheet.max_column):
                    cell = sheet.cell(row = i+1, column = j+1)
                    row.append(cell.value)
                data.append(row)
            for i in range(data.__len__()):
                if i == 0:
                    pass
                else:
                    fio = data[i][0]
                    sex = data[i][1]
                    years = data[i][2]
                    lmom = data[i][3]
                    lmo = data[i][4]
                    dta = data[i][5]
                    mta = data[i][6]
                    ogka = data[i][7]
                    gela = data[i][8]
                    sada = data[i][9]
                    dada = data[i][10]
                    hss = data[i][11]
                    hdd = data[i][12]
                    soma = data[i][13]
                    kdp = data[i][14]
                    kdl = data[i][15]
                    ik2 = mta/(dta/100*dta/100)
                    st = (100+dta+mta-160)/100
                    gi = gela*1000/mta
                    pd = sada-dada
                    dp = hss/hdd
                    udobse = 90.97+0.54*dp-0.57*dada-0.61*lmo
                    mok = udobse*hss
                    ifi = 0.011*hss+0.014*sada+0.008*dada+0.014*lmo-0.009*(dta+mta)-0.27
                    vik = 1-(dada/hss)
                    via = sada * hss
                    ir = (sada*hss)/100
                    do = gela*1000*17/100
                    mdo = hdd * do / 1000
                    si = kdp/mta
                    query = f'''
                    INSERT INTO kids (id, fio, sex, vzrg, lmom, lmo, dta, mta, ik2, ogka, st, gela, gi, sada, dada, 
                    pd, hss, dp, udobse, mok, ifi, vik, via, ir, hdd, do, mdo, soma, kdp, kdl, si)
                    VALUES ({i}, \'{fio}\', {sex}, {years}, {lmom}, {lmo}, {dta}, {mta}, {ik2}, {ogka}, {st}, {gela}, {gi},
                    {sada}, {dada}, {pd}, {hss}, {dp}, {udobse}, {mok}, {ifi}, {vik}, {via}, {ir}, {hdd}, {do},
                    {mdo}, {soma}, {kdp}, {kdl},  {si})'''
                    self.numberBox.addItem(f'{i}.{fio}')
                    try:
                        cur.execute(query)
                    except:
                        pass
            cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 1')
            boysCount = cur.fetchall()[0][0]
            cur.execute('SELECT COUNT(*) FROM kids WHERE sex = 2')
            girlsCount = cur.fetchall()[0][0]
            data = [
                ['База данных', 'успешно создана!'],
                ['',''],
                ['Мальчиков:', boysCount],
                ['Девочек:', girlsCount],
                ['',''],
            ]
            self.setData(data)
            cur.close()
            db.commit()
            db.close()
        except Exception as e:
            print(e)
            self.label.setText("Неправильное название файла")

    def onChoose(self, index):
        ind = self.numberBox.currentText().split('.')[0]
        try:
            db = sqlite3.connect(self.filename)
            cur = db.cursor()
            cur.execute(f'SELECT * FROM kids WHERE id = {ind}')
            names = list(map(lambda x: x[0], cur.description))
            for i in range(len(names)):
                names[i] = names[i].upper()
            data = list(cur.fetchall()[0])
            info = dict(zip(names, data))
            table = [
                ['ДАННЫЕ О ПАЦИЕНТЕ', ''],
                ['ФИО', info['FIO']],
                ['Возраст', info['VZRG']],
                ['Пол', 'М' if info['SEX'] == 1 else 'Ж'],
                ['', ''],
                ['АНТРОПОМЕТРИЧЕСКИЕ ПОКАЗАТЕЛИ',''],
                ['Рост', info['DTA']],
                ['Вес', info['MTA']],
                ['LMOM', info['LMOM']],
                ['LMO', info['LMO']],
                ['OGKA', info['OGKA']],
                ['Площадь тела', info['ST']],
                ['GELA', info['GELA']],
                ['SADA', info['SADA']],
                ['DADA', info['DADA']],
                ['Частота сердечных сокращений', info['HSS']],
                ['Частота дыхательных движений', info['HDD']],
                ['SOMA', info['SOMA']],
                ['KDP', info['KDP']],
                ['KDL', info['KDL']],
                ['',''],
                ['ФУНКЦИОНАЛЬНЫЕ ПОКАЗАТЕЛИ', ''],
                ['Индекс Кетла 2', info['IK2']],
                ['Жизненный индекс', info['GI']],
                ['Пульсовое давление', info['PD']],
                ['DP', info['DP']],
                ['UdObSe', info['UDOBSE']],
                ['Минутный объем кровообращения', info['MOK']],
                ['IFI', info['IFI']],
                ['Индекс Кердо', info['VIK']],
                ['Индекс Аболенской', info['VIA']],
                ['Индекс Робинсона', info['IR']],
                ['Дыхательный объем', info['DO']],
                ['Минутный дыхательный объем', info['MDO']],
                ['Силовой индекс', info['SI']]
            ]
            self.data = table
            self.setData(table)
            cur.close()
            db.close()
        except Exception as e:
            print(e)
            data = [[]]
            self.setData(data)

    def save(self):
        path = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить', '', 'Excel Files (*.xlsx)')
        wb = Workbook()
        sheet = wb.active
        i = 1
        for row in self.data:
            j = 1
            for item in row:
                cell = sheet.cell(row = i, column = j)
                cell.value = item
                j = j + 1
            j = 1
            i = i + 1
        wb.save(path[0])
        self.label.setText("Успешно сохранено!")
    
    # Добавить статистические параметры
    # # Перцентили
    # # Кросстабуляция
    # # см. фото
    # # коэфиценты кореляции между штуками
    # # графики (х - параметр 1, y - параметр 2)
    # # регрессия
    def stats(self):
        pass
            
        


def main():
    app = QtWidgets.QApplication(sys.argv)  
    app.setWindowIcon(QtGui.QIcon('../../icon.ico'))
    window = MainWindow()
    window.setWindowIcon(QtGui.QIcon('../../icon.ico'))
    window.show() 
    app.exec()  


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()